from urllib.parse import urlencode
import json

from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .backup import BackupError, create_backup_archive
from .forms import (
    BankAccountForm,
    SystemSettingsForm,
    TransactionForm,
    WalletForm,
)
from .models import BankAccount, SystemSettings, Transaction, Wallet
from .utils import (
    filtered_transactions,
    money_total,
    transaction_totals,
)


def local_app_redirect(request):
    return redirect("core:dashboard")


def dashboard(request):
    today = timezone.localdate()
    today_transactions = Transaction.objects.filter(transaction_date__date=today)
    today_incoming = today_transactions.filter(transaction_type=Transaction.CUSTOMER_PAYMENT)

    stats = {
        "incoming": money_total(today_incoming, "amount"),
        "fees": money_total(today_transactions, "fee"),
        "net": money_total(today_transactions, "net_amount"),
        "count": today_transactions.count(),
    }
    wallets = Wallet.objects.filter(is_active=True).order_by("name")
    recent_transactions = Transaction.objects.select_related("wallet", "bank_account")[:10]

    return render(
        request,
        "dashboard.html",
        {
            "stats": stats,
            "wallets": wallets,
            "recent_transactions": recent_transactions,
        },
    )


def transaction_list(request):
    transactions = filtered_transactions(request.GET)
    paginator = Paginator(transactions, 20)
    page_obj = paginator.get_page(request.GET.get("page"))
    page_params = request.GET.copy()
    page_params.pop("page", None)

    return render(
        request,
        "transactions/transaction_list.html",
        {
            "page_obj": page_obj,
            "transactions": page_obj.object_list,
            "wallets": Wallet.objects.order_by("name"),
            "banks": BankAccount.objects.order_by("bank_name", "account_name"),
            "transaction_types": Transaction.TRANSACTION_TYPES,
            "filters": request.GET,
            "page_query": page_params.urlencode(),
        },
    )


def transaction_add(request):
    form = TransactionForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        transaction = form.save(commit=False)
        transaction.created_by = None
        transaction.save()
        messages.success(request, "تم حفظ العملية بنجاح.")
        return redirect("core:transaction_list")

    return render(
        request,
        "transactions/transaction_form.html",
        {"form": form, "title": "إضافة عملية جديدة", "submit_label": "حفظ العملية"},
    )


def transaction_detail(request, pk):
    transaction = get_object_or_404(
        Transaction.objects.select_related("wallet", "bank_account"), pk=pk
    )
    return render(
        request,
        "transactions/transaction_detail.html",
        {"transaction": transaction, "can_edit": True},
    )


def transaction_edit(request, pk):
    transaction = get_object_or_404(Transaction, pk=pk)
    form = TransactionForm(request.POST or None, request.FILES or None, instance=transaction)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "تم تعديل العملية بنجاح.")
        return redirect("core:transaction_detail", pk=transaction.pk)

    return render(
        request,
        "transactions/transaction_form.html",
        {"form": form, "title": "تعديل العملية", "submit_label": "حفظ التعديل"},
    )


def transaction_delete(request, pk):
    transaction = get_object_or_404(Transaction, pk=pk)
    if request.method == "POST":
        transaction.delete()
        messages.success(request, "تم حذف العملية بنجاح.")
        return redirect("core:transaction_list")

    messages.error(request, "استخدم زر الحذف من داخل النظام لتأكيد الإجراء.")
    return redirect("core:transaction_detail", pk=transaction.pk)


def wallet_list(request):
    wallets = Wallet.objects.all()
    return render(request, "wallets/wallet_list.html", {"wallets": wallets})


def wallet_add(request):
    form = WalletForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "تمت إضافة المحفظة بنجاح.")
        return redirect("core:wallet_list")

    return render(
        request,
        "wallets/wallet_form.html",
        {"form": form, "title": "إضافة محفظة", "submit_label": "حفظ المحفظة"},
    )


def wallet_edit(request, pk):
    wallet = get_object_or_404(Wallet, pk=pk)
    form = WalletForm(request.POST or None, instance=wallet)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "تم تعديل المحفظة بنجاح.")
        return redirect("core:wallet_list")

    return render(
        request,
        "wallets/wallet_form.html",
        {"form": form, "title": "تعديل محفظة", "submit_label": "حفظ التعديل"},
    )


def wallet_delete(request, pk):
    wallet = get_object_or_404(Wallet, pk=pk)
    if request.method == "POST":
        try:
            wallet.delete()
            messages.success(request, "تم حذف المحفظة بنجاح.")
        except ProtectedError:
            messages.error(request, "لا يمكن حذف محفظة مرتبطة بعمليات مسجلة.")
    return redirect("core:wallet_list")


def bank_list(request):
    banks = BankAccount.objects.all()
    return render(request, "banks/bank_list.html", {"banks": banks})


def bank_add(request):
    form = BankAccountForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "تمت إضافة الحساب البنكي بنجاح.")
        return redirect("core:bank_list")

    return render(
        request,
        "banks/bank_form.html",
        {"form": form, "title": "إضافة حساب بنكي", "submit_label": "حفظ الحساب"},
    )


def bank_edit(request, pk):
    bank = get_object_or_404(BankAccount, pk=pk)
    form = BankAccountForm(request.POST or None, instance=bank)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "تم تعديل الحساب البنكي بنجاح.")
        return redirect("core:bank_list")

    return render(
        request,
        "banks/bank_form.html",
        {"form": form, "title": "تعديل حساب بنكي", "submit_label": "حفظ التعديل"},
    )


def bank_delete(request, pk):
    bank = get_object_or_404(BankAccount, pk=pk)
    if request.method == "POST":
        try:
            bank.delete()
            messages.success(request, "تم حذف الحساب البنكي بنجاح.")
        except ProtectedError:
            messages.error(request, "لا يمكن حذف حساب بنكي مرتبط بعمليات مسجلة.")
    return redirect("core:bank_list")


def reports(request):
    transactions = filtered_transactions(request.GET)
    totals = transaction_totals(transactions)
    export_query = urlencode(request.GET, doseq=True)

    return render(
        request,
        "reports/report.html",
        {
            "transactions": transactions,
            "totals": totals,
            "wallets": Wallet.objects.order_by("name"),
            "transaction_types": Transaction.TRANSACTION_TYPES,
            "filters": request.GET,
            "export_query": export_query,
        },
    )


def export_report_excel(request):
    transactions = filtered_transactions(request.GET)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "تقرير العمليات"
    sheet.sheet_view.rightToLeft = True

    headers = [
        "التاريخ",
        "نوع العملية",
        "المحفظة",
        "البنك",
        "العميل",
        "رقم العميل",
        "المبلغ",
        "العمولة",
        "الصافي",
        "رقم العملية",
        "ملاحظات",
    ]
    sheet.append(headers)

    for transaction in transactions:
        sheet.append(
            [
                timezone.localtime(transaction.transaction_date).strftime("%Y-%m-%d %H:%M"),
                transaction.get_transaction_type_display(),
                transaction.wallet.name,
                transaction.bank_account.bank_name if transaction.bank_account else "",
                transaction.customer_name,
                transaction.customer_phone,
                float(transaction.amount),
                float(transaction.fee),
                float(transaction.net_amount),
                transaction.reference_number,
                transaction.notes,
            ]
        )

    header_fill = PatternFill("solid", fgColor="16324F")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    widths = [18, 18, 22, 22, 22, 18, 14, 14, 14, 22, 30]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="wallet-report.xlsx"'
    workbook.save(response)
    return response


@csrf_exempt
@require_POST
def create_backup(request):
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        payload = {}

    settings_obj = SystemSettings.load()
    backup_directory = payload.get("backup_directory") or settings_obj.backup_directory

    try:
        archive_path = create_backup_archive(backup_directory)
    except BackupError as exc:
        status = 400
        code = "missing_backup_directory" if not backup_directory else "backup_failed"
        return JsonResponse({"ok": False, "code": code, "error": str(exc)}, status=status)
    except Exception as exc:
        return JsonResponse(
            {"ok": False, "code": "backup_failed", "error": str(exc)},
            status=500,
        )

    return JsonResponse(
        {
            "ok": True,
            "path": str(archive_path),
            "filename": archive_path.name,
        }
    )


def system_settings(request):
    settings_obj = SystemSettings.load()
    form = SystemSettingsForm(request.POST or None, request.FILES or None, instance=settings_obj)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "تم حفظ الإعدادات بنجاح.")
        return redirect("core:system_settings")

    return render(request, "settings/system_settings.html", {"form": form})
